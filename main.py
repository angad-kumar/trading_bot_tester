import  time
import hist_data.hist_data as hist
import technical_library.chandelier_exit as ce
import nifty_fifty_option_chain.lib as nf_data
from datetime import datetime
from openpyxl import load_workbook
import stock_analyser.stock_data as stock_value


wrkbk1 = load_workbook('data_recorder_nifty50.xlsx')
wrkbk2 = load_workbook('data_recorder_ntpc.xlsx')
# to get the active work sheet
sh1 = wrkbk1.active
sh2 = wrkbk2.active

stock1 = "^NSEI"
stock2 = "NTPC.NS"
# stock = "BTC-INR"
interval = "15m"
start_date = "2022-06-01"
end_date = "2022-06-17"
period = "1mo"

# print(stock_value.stock_hist2(stock2, interval, period))

while(1):
    row1 = sh1.max_row + 1
    row2 = sh2.max_row + 1
    
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    current_time_in_int = now.hour * 60 + now.minute
    status = ce.chandelier_exit(stock1, interval, period)
    status2 = ce.chandelier_exit(stock2, interval, period)
    
    # print(status)
    if(status == "Buy"):
        print(current_time, ":", "Bought")
        sh1.cell(row=row1, column=1).value = current_time
        sh1.cell(row=row1, column=2).value = stock_value.stock_hist2(stock1, interval, period)
        
    elif(status == "Sale"):
        print(current_time, ":" ,"Sell")
        sh1.cell(row=row1, column=1).value = current_time
        sh1.cell(row=row1, column=3).value = stock_value.stock_hist2(stock1, interval, period)
       
    if(status == "Buy"):
        print(current_time, ":", "Bought")
        sh2.cell(row=row1, column=1).value = current_time
        sh2.cell(row=row1, column=2).value = stock_value.stock_hist2(stock2, interval, period)
        
    elif(status == "Sale"):
        print(current_time, ":" ,"Sell")
        sh2.cell(row=row1, column=1).value = current_time
        sh2.cell(row=row1, column=3).value = stock_value.stock_hist2(stock2, interval, period) 
        
        
    wrkbk1.save('data_recorder_nifty50.xlsx')  
    wrkbk2.save('data_recorder_ntpc.xlsx')  
    time.sleep(60)  
    


